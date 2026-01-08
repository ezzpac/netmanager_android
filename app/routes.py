from flask import Blueprint, render_template, redirect, url_for, flash, request, jsonify, send_file
from flask_login import login_required, current_user
import json
import os
import io
import threading
import time
import zipfile
from openpyxl import Workbook, load_workbook
from .models import Device, Group, IPRange, ActivityLog, User
from . import db
from datetime import datetime
from sqlalchemy import func
from .scanner import scan_network_range

main = Blueprint('main', __name__)

import sys

def get_config_path():
    """Retorna o caminho do config.json, garantindo persistência em ambientes compilados."""
    if getattr(sys, 'frozen', False):
        # Se o app está compilado (exe), config.json fica ao lado do executável
        return os.path.join(os.path.dirname(sys.executable), 'config.json')
    # Se está em desenvolvimento, config.json fica na raiz (um nível acima de 'app/')
    return os.path.join(os.path.dirname(os.path.dirname(__file__)), 'config.json')



@main.app_context_processor
def inject_globals():
    config = {'app_name': 'NetManager', 'auto_backup': True}
    try:
        config_path = get_config_path()
        if os.path.exists(config_path):
            with open(config_path, 'r') as f:
                loaded_config = json.load(f)
                config.update(loaded_config)
    except:
        pass
        
    return dict(
        app_name=config.get('app_name', 'NetManager'),
        app_icon=config.get('app_icon', 'fas fa-network-wired'),
        auto_backup=config.get('auto_backup', True),
        network_domain=config.get('domain', ''),
        reset_requests_count=User.query.filter_by(password_reset_requested=True).count() if current_user.is_authenticated and current_user.role == 'admin' else 0
    )

def run_auto_backup():
    config_path = get_config_path()
    try:
        if os.path.exists(config_path):
            with open(config_path, 'r') as f:
                if not json.load(f).get('auto_backup', True):
                    return
    except:
        return

    # Implementation of quick JSON backup
    data = {
        'users': [{'username': u.username, 'email': u.email, 'password_hash': u.password_hash, 'role': u.role} for u in User.query.all()],
        'groups': [{'id': g.id, 'nome_grupo': g.nome_grupo} for g in Group.query.order_by(Group.nome_grupo).all()],
        'devices': [{'nome_local': d.nome_local, 'modelo': d.modelo, 'ip': d.ip, 'mac': d.mac, 'grupo_id': d.grupo_id, 'tipo': d.tipo} for d in Device.query.all()],
        'ip_ranges': [{'faixa_inicio': r.faixa_inicio, 'faixa_fim': r.faixa_fim, 'descricao': r.descricao} for r in IPRange.query.all()]
    }
    
    backup_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'backups')
    if not os.path.exists(backup_dir):
        os.makedirs(backup_dir)
        
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'auto_backup_{timestamp}.json'
    with open(os.path.join(backup_dir, filename), 'w') as f:
        json.dump(data, f, indent=4)
    print(f"Auto-backup created: {filename}")

@main.route('/')
@main.route('/dashboard')
@login_required
def dashboard():
    # Check for duplicate MAC addresses
    duplicates = db.session.query(Device.mac, func.count(Device.mac))\
        .filter(Device.mac != '', Device.mac != None)\
        .group_by(Device.mac)\
        .having(func.count(Device.mac) > 1)\
        .all()
    
    if duplicates:
        mac_list = ", ".join([mac for mac, count in duplicates])
        flash(f'Atenção: Existem dispositivos com MAC duplicado: {mac_list}. Verifique a lista de dispositivos.', 'warning')

    total_devices = Device.query.count()
    groups = Group.query.order_by(Group.nome_grupo).all()
    ip_ranges = IPRange.query.all()
    
    # Statistical data for charts
    group_stats = []
    for g in groups:
        count = Device.query.filter_by(grupo_id=g.id).count()
        group_stats.append({
            'name': g.nome_grupo, 
            'count': count,
            'icon': 'fa-layer-group',
            'color': 'info',
            'url': url_for('main.devices', group_id=g.id)
        })
    
    # Device Type Stats
    # Celular, Console, Desktop, Interruptor, LAN, Notebook, Portable, Sensor, Smart IR, Storage, Tablet, Tomada
    type_icons = {
        '3D': 'fa-cube',
        'Celular': 'fa-mobile-alt',
        'Console': 'fa-gamepad',
        'Desktop': 'fa-desktop',
        'Firewall': 'fa-shield-alt',
        'Inkjet Multifuncional': 'fa-print',
        'Interruptor': 'fa-toggle-on',
        'LAN': 'fa-network-wired',
        'Laser Multifuncional': 'fa-print',
        'Notebook': 'fa-laptop',
        'Plotter': 'fa-print',
        'Portable': 'fa-gamepad',
        'Roteador': 'fa-wifi',
        'Sensor': 'fa-microchip',
        'Servidor': 'fa-server',
        'Smart IR': 'fa-satellite-dish',
        'Smart Speaker': 'fa-volume-up',
        'Storage': 'fa-hdd',
        'Switch': 'fa-network-wired',
        'Tablet': 'fa-tablet-alt',
        'Térmica': 'fa-print',
        'Tomada': 'fa-plug'
    }
    
    # Map types to Bootstrap colors for dashboard cards
    type_colors = {
        '3D': 'dark',
        'Celular': 'primary',
        'Console': 'dark',
        'Desktop': 'info',
        'Firewall': 'danger',
        'Inkjet Multifuncional': 'dark',
        'Interruptor': 'warning',
        'LAN': 'success',
        'Laser Multifuncional': 'dark',
        'Notebook': 'info',
        'Plotter': 'dark',
        'Portable': 'warning',
        'Roteador': 'primary',
        'Sensor': 'danger',
        'Servidor': 'dark',
        'Smart IR': 'danger',
        'Smart Speaker': 'success',
        'Storage': 'primary',
        'Switch': 'success',
        'Tablet': 'primary',
        'Térmica': 'dark',
        'Tomada': 'warning'
    }

    all_types = sorted(type_icons.keys())
    type_stats = []
    for t in all_types:
        count = Device.query.filter_by(tipo=t).count()
        type_stats.append({
            'name': t, 
            'count': count, 
            'icon': type_icons.get(t, 'fa-question-circle'),
            'color': type_colors.get(t, 'primary'),
            'url': url_for('main.devices', device_type=t)
        })
        
    # Combine and Sort Stats (Total is handled separately in template)
    # Merging all groups + all device types
    combined_stats = group_stats + type_stats
    # Only show items with at least 1 device
    combined_stats = [s for s in combined_stats if s['count'] > 0]
    combined_stats.sort(key=lambda x: x['name'])
        
    # Calculate IP range usage
    ranges_with_usage = []
    all_devices_list = Device.query.all()
    
    for ipr in ip_ranges:
        count = 0
        for d in all_devices_list:
            if d.ip and '.' in d.ip:
                prefix = ".".join(d.ip.split('.')[:-1])
                try:
                    last_octet = int(d.ip.split('.')[-1])
                    if prefix == ipr.rede and ipr.faixa_inicio <= last_octet <= ipr.faixa_fim:
                        count += 1
                except ValueError:
                    continue
        
        if count > 0:
            total_in_range = ipr.faixa_fim - ipr.faixa_inicio + 1
            usage_pct = (count / total_in_range * 100) if total_in_range > 0 else 0
            ranges_with_usage.append({
                'id': ipr.id,
                'rede': ipr.rede,
                'faixa_inicio': ipr.faixa_inicio,
                'faixa_fim': ipr.faixa_fim,
                'descricao': ipr.descricao,
                'usage': min(100, usage_pct),
                'count': count,
                'total': total_in_range
            })
        
    return render_template('dashboard.html', 
                         total_devices=total_devices, 
                         group_stats=group_stats, # Kept for chart
                         combined_stats=combined_stats,
                         ip_ranges=ranges_with_usage)

@main.route('/devices')
@login_required
def devices():
    start = request.args.get('start', type=int)
    end = request.args.get('end', type=int)
    device_type = request.args.get('device_type')
    group_id = request.args.get('group_id', type=int)
    network_prefix = request.args.get('network')
    
    query = Device.query
    all_devices = query.all()
    
    # Extract unique networks (e.g., 192.168.1)
    networks = set()
    for d in all_devices:
        if d.ip and '.' in d.ip:
            prefix = ".".join(d.ip.split('.')[:-1])
            networks.add(prefix)
    networks = sorted(list(networks))
    
    filtered_devices = []
    
    # First, apply strict filters if any
    for d in all_devices:
        include = True
        
        # IP Range Filter (Last Octet)
        if start is not None and end is not None:
            if not(d.ip and '.' in d.ip):
                include = False
            else:
                try:
                    last_octet = int(d.ip.split('.')[-1])
                    if not (start <= last_octet <= end):
                        include = False
                except ValueError:
                    include = False
        
        # Network Prefix Filter (Subnet)
        if network_prefix:
            if not d.ip or not d.ip.startswith(network_prefix + '.'):
                include = False
        
        # Device Type Filter
        if device_type and device_type != '':
            if d.tipo != device_type:
                include = False
                
        # Group Filter
        if group_id is not None:
            if d.grupo_id != group_id:
                include = False
                
        if include:
            filtered_devices.append(d)
            
    devices_to_show = filtered_devices
    
    range_label = None
    if start is not None and end is not None:
        range_label = f".{start} até .{end}"
        
    groups = Group.query.order_by(Group.nome_grupo).all()
    return render_template('dispositivos.html', 
                         devices=devices_to_show, 
                         groups=groups, 
                         networks=networks,
                         active_range=range_label,
                         current_type=device_type,
                         current_group=group_id,
                         current_network=network_prefix)

@main.route('/device/add', methods=['GET', 'POST'])
@login_required
def add_device():
    if current_user.role != 'admin':
        flash('Acesso negado.', 'danger')
        return redirect(url_for('main.devices'))
        
    if request.method == 'POST':
        ip = request.form.get('ip')
        
        # Check if IP already exists
        if Device.query.filter_by(ip=ip).first():
            flash(f'Erro: O IP {ip} já está cadastrado em outro dispositivo.', 'danger')
            groups = Group.query.order_by(Group.nome_grupo).all()
            
            # Helper to preserve data
            grupo_id_str = request.form.get('grupo_id')
            grupo_id_val = int(grupo_id_str) if grupo_id_str and grupo_id_str.isdigit() else None
            
            temp_device = Device(
                nome_local=request.form.get('nome_local'),
                modelo=request.form.get('modelo'),
                ip=ip,
                mac=request.form.get('mac'),
                grupo_id=grupo_id_val,
                tipo=request.form.get('tipo'),
                observacoes=request.form.get('observacoes')
            )
            return render_template('formulario.html', groups=groups, title="Novo Dispositivo", device=temp_device)

        # Check for duplicate MAC
        mac = request.form.get('mac')
        if mac:
            mac = mac.upper()
            
        if mac and Device.query.filter_by(mac=mac).first():
            flash(f'Atenção: O endereço MAC {mac} já está associado a outro dispositivo!', 'warning')
            groups = Group.query.order_by(Group.nome_grupo).all()
            
            # Helper to preserve data (reuse temp_device logic)
            grupo_id_str = request.form.get('grupo_id')
            grupo_id_val = int(grupo_id_str) if grupo_id_str and grupo_id_str.isdigit() else None
            
            temp_device = Device(
                nome_local=request.form.get('nome_local'),
                modelo=request.form.get('modelo'),
                ip=ip,
                mac=mac,
                grupo_id=grupo_id_val,
                tipo=request.form.get('tipo'),
                observacoes=request.form.get('observacoes')
            )
            return render_template('formulario.html', groups=groups, title="Novo Dispositivo", device=temp_device)

        # Logic to add device
        new_device = Device(
            nome_local=request.form.get('nome_local'),
            modelo=request.form.get('modelo'),
            ip=ip,
            mac=request.form.get('mac').upper() if request.form.get('mac') else '',
            grupo_id=request.form.get('grupo_id') or None,
            tipo=request.form.get('tipo'),
            service_tag=request.form.get('service_tag'),
            usuario_atual=request.form.get('usuario_atual'),
            observacoes=request.form.get('observacoes')
        )
        db.session.add(new_device)
        db.session.commit()
        
        # Log activity
        log = ActivityLog(user_id=current_user.id, action=f"Adicionou dispositivo: {new_device.ip}")
        db.session.add(log)
        db.session.commit()
        
        flash('Dispositivo adicionado com sucesso!', 'success')
        return redirect(url_for('main.devices'))
        
    groups = Group.query.order_by(Group.nome_grupo).all()
    groups = Group.query.order_by(Group.nome_grupo).all()
    return render_template('formulario.html', groups=groups, title="Novo Dispositivo")

@main.route('/device/scan', methods=['POST'])
@login_required
def scan_network():
    if current_user.role != 'admin':
        flash('Acesso negado.', 'danger')
        return redirect(url_for('main.devices'))

    ip_start = request.form.get('ip_start')
    ip_end = request.form.get('ip_end')
    
    if not ip_start or not ip_end:
        flash('Faixa de IP inválida.', 'danger')
        return redirect(url_for('main.devices'))
        
    try:
        # Get domain from config
        config_path = get_config_path()
        domain = None
        if os.path.exists(config_path):
            with open(config_path, 'r') as f:
                domain = json.load(f).get('domain')

        found_devices = scan_network_range(ip_start, ip_end, domain)
        
        count_new = 0
        count_updated = 0
        
        for d_data in found_devices:
            ip = d_data['ip']
            status = d_data.get('status', False)
            
            # If host is offline, we just update status if device exists
            if not status:
                device = Device.query.filter_by(ip=ip).first()
                if device:
                    device.status = False
                    db.session.commit()
                continue

            mac = d_data.get('mac')
            hostname = d_data.get('hostname')
            vendor = d_data.get('vendor')
            service_tag = d_data.get('service_tag')
            usuario = d_data.get('usuario')
            
            # Logic:
            # 1. Identify by MAC first (if available) -> Most reliable for hardware tracking
            # 2. Identify by IP (if MAC match failed or MAC unavailable)
            
            device_by_mac = None
            if mac:
                device_by_mac = Device.query.filter_by(mac=mac).first()
                
            device_by_ip = Device.query.filter_by(ip=ip).first()
            
            if device_by_mac:
                # Device known by MAC
                if device_by_mac.ip != ip:
                    # Device moved to new IP (DHCP change)
                    # Use Case: MAC A was at IP X, now at IP Y.
                    # We must update MAC A to IP Y.
                    # Conflict Check: Is there ALREADY a device at IP Y? (device_by_ip)
                    if device_by_ip and device_by_ip.id != device_by_mac.id:
                        # Collision: IP Y is occupied by Device B.
                        # Since Device A (MAC A) is definitely at IP Y now, Device B is stale or gone.
                        # We delete Device B to free up the IP.
                        db.session.delete(device_by_ip)
                        count_updated += 1 # Technically we removed one, but effectively updating our view
                    
                    old_ip = device_by_mac.ip
                    device_by_mac.ip = ip
                    device_by_mac.data_atualizacao = datetime.utcnow()
                    count_updated += 1
                    # Log implicit? The loop commit will save changes.
                else:
                    # Device at same IP, just update update_time or other metadata
                    device_by_mac.data_atualizacao = datetime.utcnow()
                    # If model is empty, try to update with vendor info
                    if not device_by_mac.modelo and vendor:
                        device_by_mac.modelo = vendor
                    
                    device_by_mac.status = True
                    if service_tag: device_by_mac.service_tag = service_tag
                    if usuario: device_by_mac.usuario_atual = usuario
            
            elif device_by_ip:
                # Device found by IP, but MAC didn't match any existing device.
                device_by_ip.status = True
                if service_tag: device_by_ip.service_tag = service_tag
                if usuario: device_by_ip.usuario_atual = usuario

                # If device_by_ip has NO MAC, we assume it's the same device and learn the MAC.
                if not device_by_ip.mac and mac:
                    device_by_ip.mac = mac
                    if not device_by_ip.modelo and vendor:
                        device_by_ip.modelo = vendor
                    count_updated += 1
                elif device_by_ip.mac and mac and device_by_ip.mac != mac:
                    # IP Collision: IP X has Device A (MAC A) in DB, but Scan says it's Device B (MAC B).
                    # Since we couldn't find MAC B in DB (first if failed), this is a "new" device taking over an old IP.
                    # Update the existing record? Or delete and recreate?
                    # Safer to update existing record to reflect new reality if we consider "IP Slot" persistent? 
                    # No, usually Device is the Entity.
                    # Strategy: Delete old device at this IP, Create New.
                    db.session.delete(device_by_ip)
                    db.session.flush() # Ensure delete happens before insert to avoid Unique constraint checks in app memory if needed
                    
                    new_device = Device(
                        nome_local=hostname if hostname else f"Dispositivo {ip}",
                        modelo="",
                        ip=ip,
                        mac=mac,
                        grupo_id=None,
                        tipo="Outros",
                        observacoes="Adicionado via scan (substituiu dispositivo anterior no mesmo IP)"
                    )
                    db.session.add(new_device)
                    count_new += 1
            else:
                # Totally new device (New IP, New MAC)
                new_device = Device(
                    nome_local=hostname if hostname else f"Dispositivo {ip}",
                    modelo=vendor, # Use vendor as model
                    ip=ip,
                    mac=mac,
                    grupo_id=None,
                    tipo="Outros",
                    status=True,
                    service_tag=service_tag,
                    usuario_atual=usuario,
                    observacoes="Adicionado via scan de rede"
                )
                db.session.add(new_device)
                count_new += 1
        
        db.session.commit()
        
        if count_new > 0 or count_updated > 0:
            log = ActivityLog(user_id=current_user.id, action=f"Scan de rede ({ip_start}-{ip_end}): {count_new} novos, {count_updated} atualizados")
            db.session.add(log)
            db.session.commit()
            flash(f'Scan concluído! {count_new} novos dispositivos encontrados e adicionados.', 'success')
        else:
            flash('Scan concluído. Nenhum novo dispositivo encontrado.', 'info')
            
    except Exception as e:
        flash(f'Erro ao escanear rede: {str(e)}', 'danger')
        
    return redirect(url_for('main.devices'))

@main.route('/device/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_device(id):
    if current_user.role != 'admin':
        flash('Acesso negado.', 'danger')
        return redirect(url_for('main.devices'))
        
    device = Device.query.get_or_404(id)
    if request.method == 'POST':
        device.nome_local = request.form.get('nome_local')
        device.modelo = request.form.get('modelo')
        device.ip = request.form.get('ip')
        device.mac = request.form.get('mac')
        device.grupo_id = request.form.get('grupo_id') or None
        device.tipo = request.form.get('tipo')
        device.service_tag = request.form.get('service_tag')
        device.usuario_atual = request.form.get('usuario_atual')
        device.observacoes = request.form.get('observacoes')
        
        db.session.commit()
        
        # Log activity
        log = ActivityLog(user_id=current_user.id, action=f"Editou dispositivo: {device.ip}")
        db.session.add(log)
        db.session.commit()
        
        flash('Dispositivo atualizado com sucesso!', 'success')
        return redirect(url_for('main.devices'))
        
    groups = Group.query.order_by(Group.nome_grupo).all()
    return render_template('formulario.html', device=device, groups=groups, title="Editar Dispositivo")

@main.route('/device/delete/<int:id>', methods=['POST'])
@login_required
def delete_device(id):
    if current_user.role != 'admin':
        return jsonify({'success': False, 'message': 'Acesso negado'}), 403
        
    device = Device.query.get_or_404(id)
    ip_deleted = device.ip
    db.session.delete(device)
    db.session.commit()
    
    # Log activity
    log = ActivityLog(user_id=current_user.id, action=f"Excluiu dispositivo: {ip_deleted}")
    db.session.add(log)
    db.session.commit()
    
    flash('Dispositivo excluído!', 'warning')
    return redirect(url_for('main.devices'))

# --- Administrativo: Grupos ---

@main.route('/groups')
@login_required
def groups():
    if current_user.role != 'admin':
        flash('Acesso negado.', 'danger')
        return redirect(url_for('main.dashboard'))
    all_groups = Group.query.order_by(Group.nome_grupo).all()
    return render_template('groups.html', groups=all_groups)

@main.route('/group/add', methods=['POST'])
@login_required
def add_group():
    if current_user.role != 'admin':
        return jsonify({'success': False, 'message': 'Acesso negado'}), 403
    
    nome = request.form.get('nome_grupo')
    if not nome:
        flash('Nome do grupo é obrigatório.', 'danger')
        return redirect(url_for('main.groups'))
    
    new_group = Group(nome_grupo=nome)
    db.session.add(new_group)
    db.session.commit()
    
    log = ActivityLog(user_id=current_user.id, action=f"Criou grupo: {nome}")
    db.session.add(log)
    db.session.commit()
    
    flash('Grupo criado com sucesso!', 'success')
    return redirect(url_for('main.groups'))

@main.route('/group/delete/<int:id>', methods=['POST'])
@login_required
def delete_group(id):
    if current_user.role != 'admin':
        return jsonify({'success': False, 'message': 'Acesso negado'}), 403
        
    group = Group.query.get_or_404(id)
    if group.devices:
        flash('Não é possível excluir um grupo que possui dispositivos vinculados.', 'danger')
        return redirect(url_for('main.groups'))
        
    nome_deleted = group.nome_grupo
    db.session.delete(group)
    db.session.commit()
    
    log = ActivityLog(user_id=current_user.id, action=f"Excluiu grupo: {nome_deleted}")
    db.session.add(log)
    db.session.commit()
    
    flash('Grupo excluído!', 'warning')
    return redirect(url_for('main.groups'))

# --- Administrativo: Faixas de IP ---

@main.route('/ip-ranges')
@login_required
def ip_ranges():
    if current_user.role != 'admin':
        flash('Acesso negado.', 'danger')
        return redirect(url_for('main.dashboard'))
    ranges = IPRange.query.all()
    return render_template('ip_ranges.html', ranges=ranges)

@main.route('/ip-range/add', methods=['POST'])
@login_required
def add_ip_range():
    if current_user.role != 'admin':
        return jsonify({'success': False, 'message': 'Acesso negado'}), 403
    
    rede = request.form.get('rede')
    faixa_inicio = request.form.get('faixa_inicio')
    faixa_fim = request.form.get('faixa_fim')
    descricao = request.form.get('descricao')
    
    new_range = IPRange(rede=rede, faixa_inicio=int(faixa_inicio), faixa_fim=int(faixa_fim), descricao=descricao)
    db.session.add(new_range)
    db.session.commit()
    
    log = ActivityLog(user_id=current_user.id, action=f"Criou faixa de IP: {rede}.{faixa_inicio}-{faixa_fim}")
    db.session.add(log)
    db.session.commit()
    
    flash('Faixa de IP cadastrada com sucesso!', 'success')
    return redirect(url_for('main.ip_ranges'))

@main.route('/ip-range/edit/<int:id>', methods=['POST'])
@login_required
def edit_ip_range(id):
    if current_user.role != 'admin':
        return jsonify({'success': False, 'message': 'Acesso negado'}), 403
        
    ipr = IPRange.query.get_or_404(id)
    
    ipr.rede = request.form.get('rede')
    ipr.faixa_inicio = int(request.form.get('faixa_inicio'))
    ipr.faixa_fim = int(request.form.get('faixa_fim'))
    ipr.descricao = request.form.get('descricao')
    
    db.session.commit()
    
    log = ActivityLog(user_id=current_user.id, action=f"Editou faixa de IP: {ipr.rede}.{ipr.faixa_inicio}-{ipr.faixa_fim}")
    db.session.add(log)
    db.session.commit()
    
    flash('Faixa de IP atualizada com sucesso!', 'success')
    return redirect(url_for('main.ip_ranges'))

@main.route('/ip-range/delete/<int:id>', methods=['POST'])
@login_required
def delete_ip_range(id):
    if current_user.role != 'admin':
        return jsonify({'success': False, 'message': 'Acesso negado'}), 403
        
    ipr = IPRange.query.get_or_404(id)
    desc = ipr.descricao
    full_range = f"{ipr.rede}.{ipr.faixa_inicio}-{ipr.faixa_fim}"
    
    db.session.delete(ipr)
    db.session.commit()
    
    log = ActivityLog(user_id=current_user.id, action=f"Excluiu faixa de IP: {full_range} ({desc})")
    db.session.add(log)
    db.session.commit()
    
    flash('Faixa de IP excluída!', 'warning')
    return redirect(url_for('main.ip_ranges'))

@main.route('/ip-range/clear-devices/<int:id>', methods=['POST'])
@login_required
def clear_ip_range_devices(id):
    if current_user.role != 'admin':
        flash('Acesso negado.', 'danger')
        return redirect(url_for('main.ip_ranges'))
        
    ipr = IPRange.query.get_or_404(id)
    all_devices = Device.query.all()
    deleted_count = 0
    
    for d in all_devices:
        if d.ip and '.' in d.ip:
            prefix = ".".join(d.ip.split('.')[:-1])
            try:
                last_octet = int(d.ip.split('.')[-1])
                if prefix == ipr.rede and ipr.faixa_inicio <= last_octet <= ipr.faixa_fim:
                    db.session.delete(d)
                    deleted_count += 1
            except ValueError:
                continue
    
    if deleted_count > 0:
        db.session.commit()
        log = ActivityLog(user_id=current_user.id, action=f"Limpou {deleted_count} dispositivos da faixa: {ipr.rede}.{ipr.faixa_inicio}-{ipr.faixa_fim}")
        db.session.add(log)
        db.session.commit()
        flash(f'{deleted_count} dispositivos foram excluídos desta faixa.', 'success')
    else:
        flash('Nenhum dispositivo encontrado nesta faixa para excluir.', 'info')
        
    return redirect(url_for('main.ip_ranges'))

# --- Administrativo: Logs e Usuários ---

@main.route('/logs')
@login_required
def view_logs():
    if current_user.role != 'admin':
        flash('Acesso negado.', 'danger')
        return redirect(url_for('main.dashboard'))
    logs = ActivityLog.query.order_by(ActivityLog.timestamp.desc()).limit(200).all()
    return render_template('logs.html', logs=logs)

@main.route('/users')
@login_required
def users():
    if current_user.role != 'admin':
        flash('Acesso negado.', 'danger')
        return redirect(url_for('main.dashboard'))
    all_users = User.query.all()
    return render_template('users.html', users=all_users)

@main.route('/user/add', methods=['POST'])
@login_required
def add_user():
    if current_user.role != 'admin':
        return jsonify({'success': False, 'message': 'Acesso negado'}), 403
    
    username = request.form.get('username')
    email = request.form.get('email')
    password = request.form.get('password')
    role = request.form.get('role', 'user')
    
    if User.query.filter_by(username=username).first():
        flash('Usuário já existe.', 'danger')
        return redirect(url_for('main.users'))
        
    if User.query.filter_by(email=email).first():
        flash('E-mail já está em uso.', 'danger')
        return redirect(url_for('main.users'))
        
    new_user = User(username=username, email=email, role=role)
    new_user.set_password(password)
    db.session.add(new_user)
    db.session.commit()
    
    flash('Usuário criado com sucesso!', 'success')
    return redirect(url_for('main.users'))

@main.route('/user/delete/<int:id>', methods=['POST'])
@login_required
def delete_user(id):
    if current_user.role != 'admin':
        flash('Acesso negado.', 'danger')
        return redirect(url_for('main.users'))

    user = User.query.get_or_404(id)

    if user.id == current_user.id:
        flash('Você não pode excluir sua própria conta.', 'danger')
        return redirect(url_for('main.users'))

    username = user.username
    db.session.delete(user)
    db.session.commit()

    log = ActivityLog(
        user_id=current_user.id,
        action=f"Excluiu permanentemente o usuário: {username}"
    )
    db.session.add(log)
    db.session.commit()

    flash(f'Usuário {username} excluído com sucesso.', 'success')
    return redirect(url_for('main.users'))

@main.route('/user/toggle-status/<int:id>', methods=['POST'])
@login_required
def toggle_user_status(id):
    if current_user.role != 'admin':
        flash('Acesso negado.', 'danger')
        return redirect(url_for('main.users'))

    user = User.query.get_or_404(id)

    if user.id == current_user.id:
        flash('Você não pode desativar seu próprio usuário.', 'warning')
        return redirect(url_for('main.users'))

    user.active = not user.active
    db.session.commit()

    status = 'ativado' if user.active else 'desativado'
    log = ActivityLog(
        user_id=current_user.id,
        action=f"{status.capitalize()} usuário: {user.username}"
    )
    db.session.add(log)
    db.session.commit()

    flash(f'Usuário {status} com sucesso.', 'success')
    return redirect(url_for('main.users'))

@main.route('/user/reset-password/<int:id>', methods=['POST'])
@login_required
def reset_user_password(id):
    if current_user.role != 'admin':
        flash('Acesso negado.', 'danger')
        return redirect(url_for('main.users'))

    user = User.query.get_or_404(id)

    if user.id == current_user.id:
        flash('Você não pode resetar sua própria senha por aqui.', 'warning')
        return redirect(url_for('main.users'))

    temp_password = '123456'
    user.set_password(temp_password)
    user.password_reset_requested = False
    db.session.commit()

    log = ActivityLog(
        user_id=current_user.id,
        action=f"Resetou senha do usuário: {user.username}"
    )
    db.session.add(log)
    db.session.commit()

    flash(f'Senha do usuário resetada. Nova senha: {temp_password}', 'warning')
    return redirect(url_for('main.users'))

@main.route('/user/change-role/<int:id>', methods=['POST'])
@login_required
def change_user_role(id):
    if current_user.role != 'admin':
        flash('Acesso negado.', 'danger')
        return redirect(url_for('main.users'))

    user = User.query.get_or_404(id)

    if user.id == current_user.id:
        flash('Você não pode alterar seu próprio perfil.', 'warning')
        return redirect(url_for('main.users'))

    new_role = request.form.get('role')

    if new_role not in ['admin', 'user']:
        flash('Perfil inválido.', 'danger')
        return redirect(url_for('main.users'))

    old_role = user.role
    user.role = new_role
    db.session.commit()

    log = ActivityLog(
        user_id=current_user.id,
        action=f"Alterou perfil do usuário {user.username}: {old_role} → {new_role}"
    )
    db.session.add(log)
    db.session.commit()

    flash(f'Perfil do usuário {user.username} atualizado para {new_role}.', 'success')
    return redirect(url_for('main.users'))

@main.route('/user/change-email/<int:id>', methods=['POST'])
@login_required
def change_user_email(id):
    if current_user.role != 'admin':
        flash('Acesso negado.', 'danger')
        return redirect(url_for('main.users'))

    user = User.query.get_or_404(id)
    new_email = request.form.get('email')

    if not new_email:
        flash('E-mail é obrigatório.', 'danger')
        return redirect(url_for('main.users'))

    existing_user = User.query.filter_by(email=new_email).first()
    if existing_user and existing_user.id != user.id:
        flash('E-mail já está em uso por outro usuário.', 'danger')
        return redirect(url_for('main.users'))

    old_email = user.email
    user.email = new_email
    db.session.commit()

    log = ActivityLog(
        user_id=current_user.id,
        action=f"Alterou e-mail do usuário {user.username}: {old_email} → {new_email}"
    )
    db.session.add(log)
    db.session.commit()

    flash(f'E-mail do usuário {user.username} atualizado com sucesso.', 'success')
    return redirect(url_for('main.users'))

# --- Configurações, Backup e Restauração ---

@main.route('/settings')
@login_required
def settings():
    if current_user.role != 'admin':
        flash('Acesso negado.', 'danger')
        return redirect(url_for('main.dashboard'))
    
    # Get current port from config
    port = 5000
    try:
        config_path = get_config_path()
        if os.path.exists(config_path):
            with open(config_path, 'r') as f:
                port = json.load(f).get('port', 5000)
    except:
        pass
        
    return render_template('settings.html', current_port=port)

@main.route('/settings/save-port', methods=['POST'])
@login_required
def save_port():
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
        
    port = request.form.get('port')
    if not port or not port.isdigit():
        flash('Porta inválida.', 'danger')
        return redirect(url_for('main.settings'))
        
    try:
        config_path = get_config_path()
        config = {}
        if os.path.exists(config_path):
            with open(config_path, 'r') as f:
                config = json.load(f)
        
        config['port'] = int(port)
        with open(config_path, 'w') as f:
            json.dump(config, f, indent=4)
            
        flash(f'Porta alterada para {port}. Por favor, reinicie o programa para aplicar.', 'success')
    except Exception as e:
        flash(f'Erro ao salvar configuração: {str(e)}', 'danger')
        
    return redirect(url_for('main.settings'))

@main.route('/settings/save-visual', methods=['POST'])
@login_required
def save_visual():
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
        
    app_name = request.form.get('app_name')
    if not app_name:
        flash('Nome do sistema inválido.', 'danger')
        return redirect(url_for('main.settings'))
        
    try:
        config_path = get_config_path()
        config = {}
        if os.path.exists(config_path):
            with open(config_path, 'r') as f:
                config = json.load(f)
        
        config['app_name'] = app_name
        with open(config_path, 'w') as f:
            json.dump(config, f, indent=4)
            
        flash(f'Nome do sistema atualizado para "{app_name}".', 'success')
    except Exception as e:
        flash(f'Erro ao salvar configuração visual: {str(e)}', 'danger')
        
    return redirect(url_for('main.settings'))

@main.route('/settings/toggle-auto-backup', methods=['POST'])
@login_required
def toggle_auto_backup():
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
        
    try:
        config_path = get_config_path()
        config = {}
        if os.path.exists(config_path):
            with open(config_path, 'r') as f:
                config = json.load(f)
        
        config['auto_backup'] = request.form.get('auto_backup') == 'on'
        with open(config_path, 'w') as f:
            json.dump(config, f, indent=4)
            
        status = 'ativado' if config['auto_backup'] else 'desativado'
        flash(f'Backup automático {status}.', 'success')
    except Exception as e:
        flash(f'Erro ao alternar backup automático: {str(e)}', 'danger')
        
    return redirect(url_for('main.settings'))

@main.route('/settings/save-identity', methods=['POST'])
@login_required
def save_identity():
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
        
    app_icon = request.form.get('app_icon')
    
    try:
        config_path = get_config_path()
        config = {}
        if os.path.exists(config_path):
            with open(config_path, 'r') as f:
                config = json.load(f)
        
        if app_icon: config['app_icon'] = app_icon
        
        with open(config_path, 'w') as f:
            json.dump(config, f, indent=4)
            
        flash('Identidade visual atualizada!', 'success')
    except Exception as e:
        flash(f'Erro ao salvar identidade visual: {str(e)}', 'danger')
        
    return redirect(url_for('main.settings'))

@main.route('/settings/save-domain', methods=['POST'])
@login_required
def save_domain():
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
        
    domain = request.form.get('domain')
    
    try:
        config_path = get_config_path()
        config = {}
        if os.path.exists(config_path):
            with open(config_path, 'r') as f:
                config = json.load(f)
        
        config['domain'] = domain
        
        with open(config_path, 'w') as f:
            json.dump(config, f, indent=4)
            
        flash(f'Domínio da rede atualizado para "{domain}".', 'success')
    except Exception as e:
        flash(f'Erro ao salvar domínio: {str(e)}', 'danger')
        
    return redirect(url_for('main.settings'))

@main.route('/settings/clear-logs', methods=['POST'])
@login_required
def clear_logs():
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
        
    try:
        ActivityLog.query.delete()
        db.session.commit()
        flash('Todos os logs de atividade foram excluídos com sucesso.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir logs: {str(e)}', 'danger')
        
    return redirect(url_for('main.settings'))

@main.route('/settings/reset-all', methods=['POST'])
@login_required
def reset_all():
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
        
    try:
        # Delete all data
        Device.query.delete()
        IPRange.query.delete()
        Group.query.delete()
        ActivityLog.query.delete()
        
        # Log the reset action
        log = ActivityLog(user_id=current_user.id, action="Realizou exclusão total de dados do sistema")
        db.session.add(log)
        db.session.commit()
        
        flash('Todas as informações foram excluídas com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir informações: {str(e)}', 'danger')
        
    return redirect(url_for('main.settings'))

@main.route('/settings/backup')
@login_required
def backup_database():
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    
    # 1. Gather all database data
    data = {
        'users': [],
        'groups': [],
        'devices': [],
        'ip_ranges': [],
        'logs': []
    }
    
    for user in User.query.all():
        data['users'].append({
            'username': user.username,
            'email': user.email,
            'password_hash': user.password_hash,
            'role': user.role,
            'active': user.active
        })
        
    for group in Group.query.order_by(Group.nome_grupo).all():
        data['groups'].append({
            'id': group.id,
            'nome_grupo': group.nome_grupo
        })
        
    for device in Device.query.all():
        data['devices'].append({
            'nome_local': device.nome_local,
            'modelo': device.modelo,
            'ip': device.ip,
            'mac': device.mac,
            'grupo_id': device.grupo_id,
            'tipo': device.tipo,
            'status': device.status,
            'service_tag': device.service_tag,
            'usuario_atual': device.usuario_atual,
            'observacoes': device.observacoes
        })
        
    for ip_range in IPRange.query.all():
        data['ip_ranges'].append({
            'rede': ip_range.rede,
            'faixa_inicio': ip_range.faixa_inicio,
            'faixa_fim': ip_range.faixa_fim,
            'descricao': ip_range.descricao
        })

    for log in ActivityLog.query.all():
        data['logs'].append({
            'user_id': log.user_id,
            'action': log.action,
            'timestamp': log.timestamp.isoformat() if log.timestamp else None
        })
        
    # 2. Create JSON in memory
    json_data = json.dumps(data, indent=4)
    
    # 3. Create ZIP in memory
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        # Add database data as a JSON file
        zf.writestr('database_backup.json', json_data)
        
        # Add current config.json if it exists
        config_path = get_config_path()
        if os.path.exists(config_path):
            with open(config_path, 'r') as f:
                zf.writestr('config.json', f.read())
    
    zip_buffer.seek(0)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    return send_file(
        zip_buffer,
        as_attachment=True,
        download_name=f'full_backup_netmanager_{timestamp}.zip',
        mimetype='application/zip'
    )

@main.route('/settings/restore', methods=['POST'])
@login_required
def restore_database():
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
        
    file = request.files.get('backup_file')
    if not file:
        flash('Nenhum arquivo enviado.', 'danger')
        return redirect(url_for('main.settings'))
        
    filename = file.filename.lower()
    
    # Get restore selections from form
    restore_users = request.form.get('restore_users') == 'on'
    restore_groups = request.form.get('restore_groups') == 'on'
    restore_devices = request.form.get('restore_devices') == 'on'
    restore_ip_ranges = request.form.get('restore_ip_ranges') == 'on'
    restore_logs = request.form.get('restore_logs') == 'on'
    restore_config = request.form.get('restore_config') == 'on'
    
    # If nothing selected, restore everything (backward compatibility)
    if not any([restore_users, restore_groups, restore_devices, restore_ip_ranges, restore_logs, restore_config]):
        restore_users = restore_groups = restore_devices = True
        restore_ip_ranges = restore_logs = restore_config = True
    
    try:
        run_auto_backup() # Automatic backup before restoration
        
        data = None
        config_content = None
        
        if filename.endswith('.zip'):
            with zipfile.ZipFile(file, 'r') as zf:
                if 'database_backup.json' in zf.namelist():
                    with zf.open('database_backup.json') as f:
                        data = json.load(f)
                if 'config.json' in zf.namelist():
                    with zf.open('config.json') as f:
                        config_content = f.read().decode('utf-8')
        elif filename.endswith('.json'):
            data = json.load(file)
        else:
            flash('Formato de arquivo não suportado. Use .json ou .zip', 'warning')
            return redirect(url_for('main.settings'))
            
        if not data:
            flash('Dados da base não encontrados no arquivo.', 'danger')
            return redirect(url_for('main.settings'))

        # 1. Delete only selected tables
        if restore_devices:
            Device.query.delete()
        if restore_ip_ranges:
            IPRange.query.delete()
        if restore_groups:
            Group.query.delete()
        if restore_logs:
            ActivityLog.query.delete()
        
        # 2. Restore only selected data
        
        # Restore Groups first (dependencies)
        if restore_groups:
            for g_data in data.get('groups', []):
                group = Group(id=g_data['id'], nome_grupo=g_data['nome_grupo'])
                db.session.add(group)
            db.session.flush()
        
        # Restore IP Ranges
        if restore_ip_ranges:
            for r_data in data.get('ip_ranges', []):
                ipr = IPRange(
                    rede=r_data.get('rede', '192.168.0'),
                    faixa_inicio=r_data['faixa_inicio'],
                    faixa_fim=r_data['faixa_fim'],
                    descricao=r_data.get('descricao', '')
                )
                db.session.add(ipr)
            
        # Restore Devices
        if restore_devices:
            for d_data in data.get('devices', []):
                device = Device(**d_data)
                db.session.add(device)

        # Restore Logs
        if restore_logs:
            for l_data in data.get('logs', []):
                log = ActivityLog(
                    user_id=l_data.get('user_id'),
                    action=l_data.get('action'),
                    timestamp=datetime.fromisoformat(l_data['timestamp']) if l_data.get('timestamp') else datetime.utcnow()
                )
                db.session.add(log)
            
        # Restore Users (update or create)
        if restore_users:
            for u_data in data.get('users', []):
                existing_user = User.query.filter_by(username=u_data['username']).first()
                if existing_user:
                    # Update existing user with backup data
                    existing_user.email = u_data.get('email')
                    existing_user.password_hash = u_data['password_hash']
                    existing_user.role = u_data.get('role', 'user')
                    existing_user.active = u_data.get('active', True)
                else:
                    # Create new user from backup
                    user = User(
                        username=u_data['username'],
                        email=u_data.get('email'),
                        password_hash=u_data['password_hash'],
                        role=u_data.get('role', 'user'),
                        active=u_data.get('active', True)
                    )
                    db.session.add(user)
                
        # Restore config.json if selected
        if restore_config and config_content:
            config_path = get_config_path()
            try:
                # Validate JSON before writing
                json.loads(config_content) 
                with open(config_path, 'w') as f:
                    f.write(config_content)
            except:
                pass # Skip if invalid

        db.session.commit()
        
        # Build list of restored items for feedback
        restored_items = []
        if restore_users: restored_items.append('usuários')
        if restore_groups: restored_items.append('grupos')
        if restore_devices: restored_items.append('dispositivos')
        if restore_ip_ranges: restored_items.append('faixas de IP')
        if restore_logs: restored_items.append('logs')
        if restore_config: restored_items.append('configurações')
        
        log = ActivityLog(user_id=current_user.id, action=f"Restaurou backup: {', '.join(restored_items)}")
        db.session.add(log)
        db.session.commit()
        
        flash(f'Restauração concluída com sucesso: {", ".join(restored_items)}', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao restaurar: {str(e)}', 'danger')
        
    return redirect(url_for('main.settings'))

@main.route('/settings/export-excel')
@login_required
def export_excel():
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from collections import defaultdict
    import ipaddress
    
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
        
    devices = Device.query.all()
    
    # Function to extract IP network prefix (first 3 octets)
    def get_ip_prefix(ip_str):
        if not ip_str or '.' not in ip_str:
            return 'Sem IP'
        parts = ip_str.split('.')
        if len(parts) >= 3:
            return f"{parts[0]}.{parts[1]}.{parts[2]}"
        return 'Sem IP'
    
    # Function to sort IPs naturally
    def ip_sort_key(device):
        try:
            if device.ip and '.' in device.ip:
                return ipaddress.IPv4Address(device.ip)
            return ipaddress.IPv4Address('255.255.255.255')
        except:
            return ipaddress.IPv4Address('255.255.255.255')
    
    # Sort devices by IP address
    sorted_devices = sorted(devices, key=ip_sort_key)
    
    # Group devices by IP range
    ip_ranges = defaultdict(list)
    for d in sorted_devices:
        prefix = get_ip_prefix(d.ip)
        ip_ranges[prefix].append(d)
    
    # Sort IP ranges
    sorted_ranges = sorted(ip_ranges.keys(), key=lambda x: 
        tuple(int(p) for p in x.split('.')) if x != 'Sem IP' else (999, 999, 999))
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Dispositivos'
    
    # Define styles
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    range_header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    range_header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    range_header_alignment = Alignment(horizontal='left', vertical='center')
    
    cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
    
    odd_row_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    even_row_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    
    # Headers
    headers = ['Status', 'Usuário', 'Hostname', 'Modelo', 'Service Tag', 'IP', 'MAC', 'Grupo', 'Tipo', 'Observações', 'Data Criação']
    ws.append(headers)
    
    # Style header row
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Set column widths
    ws.column_dimensions['A'].width = 10  # Status
    ws.column_dimensions['B'].width = 25  # Usuário
    ws.column_dimensions['C'].width = 20  # Hostname
    ws.column_dimensions['D'].width = 20  # Modelo
    ws.column_dimensions['E'].width = 15  # Service Tag
    ws.column_dimensions['F'].width = 16  # IP
    ws.column_dimensions['G'].width = 18  # MAC
    ws.column_dimensions['H'].width = 15  # Grupo
    ws.column_dimensions['I'].width = 12  # Tipo
    ws.column_dimensions['J'].width = 35  # Observações
    ws.column_dimensions['K'].width = 18  # Data Criação
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    current_row = 2
    
    # Add devices grouped by IP range
    for range_idx, ip_range in enumerate(sorted_ranges):
        # Add range header row
        ws.merge_cells(f'A{current_row}:K{current_row}')
        range_cell = ws.cell(row=current_row, column=1)
        range_cell.value = f'Faixa de IP: {ip_range}.0/24'
        range_cell.font = range_header_font
        range_cell.fill = range_header_fill
        range_cell.alignment = range_header_alignment
        
        # Apply border to all cells in the merged range
        for col in range(1, 12):
            cell = ws.cell(row=current_row, column=col)
            cell.border = thin_border
        
        current_row += 1
        
        # Add devices in this range
        devices_in_range = ip_ranges[ip_range]
        for device_idx, d in enumerate(devices_in_range):
            row_data = [
                'Online' if d.status else 'Offline',
                d.usuario_atual,
                d.nome_local,
                d.modelo,
                d.service_tag,
                d.ip,
                d.mac,
                d.group.nome_grupo if d.group else 'N/A',
                d.tipo,
                d.observacoes,
                d.data_criacao.strftime('%d/%m/%Y %H:%M') if d.data_criacao else ''
            ]
            ws.append(row_data)
            
            # Apply styling to data row
            row_fill = odd_row_fill if device_idx % 2 == 0 else even_row_fill
            
            for col_num in range(1, 12):
                cell = ws.cell(row=current_row, column=col_num)
                cell.alignment = cell_alignment
                cell.fill = row_fill
                cell.border = thin_border
                
                # Special formatting for specific columns
                if col_num == 5:  # IP column
                    cell.font = Font(name='Consolas', size=10)
                elif col_num == 6:  # MAC column
                    cell.font = Font(name='Consolas', size=10)
            
            current_row += 1
        
        # Add blank row between ranges (except after last range)
        if range_idx < len(sorted_ranges) - 1:
            current_row += 1
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    return send_file(
        output,
        as_attachment=True,
        download_name=f'dispositivos_{timestamp}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@main.route('/settings/import-excel', methods=['POST'])
@login_required
def import_excel():
    if current_user.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
        
    file = request.files.get('excel_file')
    if not file:
        flash('Nenhum arquivo enviado.', 'danger')
        return redirect(url_for('main.settings'))
        
    try:
        run_auto_backup() # Automatic backup before mass import
        wb = load_workbook(file)
        ws = wb.active
        
        # We assume headers are in the first row and map columns by name
        headers = [str(cell.value).strip() for cell in ws[1]]
        
        def get_val(row_cells, col_name):
            try:
                idx = headers.index(col_name)
                return row_cells[idx].value
            except (ValueError, IndexError):
                return None

        count_added = 0
        count_updated = 0
        
        # Iterate from row 2
        for row_cells in ws.iter_rows(min_row=2):
            ip = str(get_val(row_cells, 'IP') or '').strip()
            if not ip or ip.lower() == 'nan' or ip == 'None':
                continue
                
            group_name = str(get_val(row_cells, 'Grupo') or 'Geral').strip()
            group = Group.query.filter_by(nome_grupo=group_name).first()
            if not group:
                group = Group(nome_grupo=group_name)
                db.session.add(group)
                db.session.flush()
                
            device = Device.query.filter_by(ip=ip).first()
            
            # Get values with new header names, fallback to old ones
            v_hostname = get_val(row_cells, 'Hostname') or get_val(row_cells, 'Nome Local')
            v_modelo = get_val(row_cells, 'Modelo')
            v_mac = get_val(row_cells, 'MAC')
            v_tipo = get_val(row_cells, 'Tipo')
            v_obs = get_val(row_cells, 'Observações')
            v_usuario = get_val(row_cells, 'Usuário')
            v_service_tag = get_val(row_cells, 'Service Tag')

            if device:
                # Update existing device
                if v_hostname: device.nome_local = str(v_hostname)
                if v_modelo is not None: device.modelo = str(v_modelo)
                if v_mac is not None: device.mac = str(v_mac)
                if v_tipo: device.tipo = str(v_tipo)
                if v_obs is not None: device.observacoes = str(v_obs)
                if v_usuario is not None: device.usuario_atual = str(v_usuario)
                if v_service_tag is not None: device.service_tag = str(v_service_tag)
                device.grupo_id = group.id
                count_updated += 1
            else:
                # Create new device
                new_device = Device(
                    nome_local=str(v_hostname or 'Desconhecido'),
                    modelo=str(v_modelo or ''),
                    ip=ip,
                    mac=str(v_mac or ''),
                    grupo_id=group.id,
                    tipo=str(v_tipo or 'Desktop'),
                    usuario_atual=str(v_usuario or ''),
                    service_tag=str(v_service_tag or ''),
                    observacoes=str(v_obs or '')
                )
                db.session.add(new_device)
                count_added += 1
                
        db.session.commit()
        
        log = ActivityLog(user_id=current_user.id, action=f"Importou Excel: {count_added} novos, {count_updated} atualizados")
        db.session.add(log)
        db.session.commit()
        
        flash(f'Importação concluída: {count_added} adicionados, {count_updated} atualizados.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao importar Excel: {str(e)}', 'danger')
        
    return redirect(url_for('main.settings'))


@main.route("/shutdown", methods=["POST"])
@login_required
def api_shutdown():
    """Encerra o app de forma limpa"""
    if current_user.role != 'admin':
        return jsonify({"message": "Acesso negado"}), 403

    def stop():
        # Aguarda 1.5 segundos para garantir que a resposta JSON 
        # e o redirecionamento no front-end aconteçam
        time.sleep(1.5)  
        
        # Encerramento robusto para Windows (PyInstaller EXE)
        if os.name == 'nt':
            import subprocess
            # /F - Forçar encerramento
            # /T - Encerrar processos filhos (árvore de processos)
            # /PID - ID do processo atual
            subprocess.run(['taskkill', '/F', '/T', '/PID', str(os.getpid())], 
                           capture_output=True, creationflags=subprocess.CREATE_NO_WINDOW)
        else:
            os._exit(0) # Fallback para outros sistemas

    # Executa o encerramento em uma thread separada para não travar a requisição atual
    threading.Thread(target=stop).start()
    
